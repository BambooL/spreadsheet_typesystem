import openpyxl
import parser
import os
import re
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.formula import Tokenizer
from openpyxl.styles import Color, Fill, PatternFill
from openpyxl.cell import Cell
import headerchecker


class Header_cell:
	def __init__(self, cell_type, cell_value):
		self.header = []
		self.type = cell_type
		self.value = cell_value

	def add_header(self, headercell):
		self.header.append(headercell)

	def add_OR_header(self, cell_range):
		start = cell_range.split(":")[0]
		end = cell_range.split(":")[1]
		start_letter = re.findall('[a-zA-Z]+', start)[0]
		start_digit = re.findall('\d+', start)[0]
		end_letter = re.findall('[a-zA-Z]+', end)[0]
		end_digit = re.findall('\d+', end)[0]
		or_type = []
		if (int(end_digit) - int(start_digit) != 0 and start_letter == end_letter):
			for i in range(int(end_digit) - int(start_digit) + 1):
				or_type.append(start_letter + str(int(start_digit) + i))
		self.header.append(or_type) 

def headerspread(cell, sheet, store_header):
	if (column_data(cell, sheet) > row_data(cell, sheet)):
		for sub_cell in following_column(cell, sheet):
			store_header = update1(cell, sub_cell, store_header)
	else:
		for sub_cell in following_row(cell, sheet):
			store_header = update1(cell, sub_cell, store_header)
	return store_header

def column_data(cell, sheet):
	n = 0
	for i in range(cell.row + 1, sheet.max_row + 1):
		coordinate = cell.column + str(i)
		if (sheet[coordinate].data_type == "n" and sheet[coordinate].value != None):
			n = n + 1
	return n

def row_data(cell, sheet):
	n = 0
	for i in range(ord(cell.column) - ord("A") + 1, sheet.max_column):
		coordinate = chr(i + ord("A")) + str(cell.row)
		if (sheet[coordinate].data_type == "n" and sheet[coordinate].value != None):
			n = n + 1
	return n


def update1(cell, sub_cell, store_header):
	if sub_cell.coordinate in store_header.keys():
		store_header[sub_cell.coordinate].append(cell.coordinate)
		if (cell.coordinate in store_header.keys()):
			for item in store_header[cell.coordinate]:
				if (type(item) != list and item in store_header[sub_cell.coordinate] and is_header(sheet[item])):
					store_header[sub_cell.coordinate].remove(item)

	else:
		header = []
		header.append(cell.coordinate)
		store_header[sub_cell.coordinate] = header
	return store_header

def update2(cell, refrange, store_header):
	if cell.coordinate in store_header.keys():
		store_header[cell.coordinate].append(add_OR_header(refrange))
	else:
		header = []
		header.append(add_OR_header(refrange))
		store_header[cell.coordinate] = header
	return store_header

def add_OR_header(cell_range):
	start = cell_range.split(":")[0]
	end = cell_range.split(":")[1]
	start_letter = re.findall('[a-zA-Z]+', start)[0]
	start_digit = re.findall('\d+', start)[0]
	end_letter = re.findall('[a-zA-Z]+', end)[0]
	end_digit = re.findall('\d+', end)[0]
	or_type = []
	if (int(end_digit) - int(start_digit) != 0 and start_letter == end_letter):
		for i in range(int(end_digit) - int(start_digit) + 1):
			or_type.append(start_letter + str(int(start_digit) + i))
	return or_type

def following_column(cell, sheet):
	column_cell = []
	column = cell.column
	next_cell = sheet[column + str(cell.row + 1)]
	for i in range(cell.row + 2, sheet.max_row + 1):
		coordinate = column + str(i)
		cur_cell = next_cell
		next_cell = sheet[coordinate]
		# if a blank cell followed by a headercell in this column, stop the headerspreading
		if (cur_cell.value == None and next_cell.data_type == 's'):
			break
		else:
			if (cur_cell.value != None):
				column_cell.append(cur_cell)
	return column_cell


def following_row(cell, sheet):
	row_cell = []
	row = cell.row
	next_cell = sheet[chr(ord(cell.column) + 1) + str(row)]
	for i in range(ord(cell.column) - ord("A") + 2, sheet.max_column):
		coordinate = chr(i + ord("A")) + str(row)
		cur_cell = next_cell
		next_cell = sheet[coordinate]
		if (cur_cell.value == None and next_cell.data_type == 's'):
			break
		else:
			if (cur_cell.value != None):
				row_cell.append(cur_cell)
	return row_cell


def is_header(cell):
	if (cell.data_type == 's' and cell.style.font.bold):
		return True
	else:
		return False

if __name__ == "__main__":
	# rootdir ='../data/alldata/'
	# loop inside the workbook
	# for files in os.walk(rootdir):
	# 	print type(files)
	# 	folder = files[2]
	# 	for i in range(len(folder)):
	# 		if (re.search("xlsx", str(folder[i]))):
	wb = load_workbook(filename = "../data/cs101/bad/test.xlsx")
	sheet = wb["Sheet1"]
	store_header = {}
	problem_cell = []
	for row in sheet:
		for cell in row:
			if (is_header(cell)):
				if (cell.coordinate in store_header.keys()):
					store_header[cell.coordinate].append("1")
				else:
					store_header[cell.coordinate] = ["1"]
				store_header = headerspread(cell, sheet, store_header)
				for item in sheet._merged_cells:
					if (item.find(cell.coordinate) != -1):
						for r in sheet[item]:
							for c in r:
								store_header[c.coordinate] = ["1"]
								store_header = headerspread(c, sheet, store_header)
			elif (cell.data_type == 'f'):
				tok = Tokenizer(cell._value)
				tok.parse()
				for token in tok.items:
					if (re.match(r'[A-Z][0-9]+$', cell.value[1:])):
						store_header = update1(sheet[cell.value[1:]], cell, store_header)
					if (token.value == "AVERAGE(" or token.value == "SUM("):
						refrange = re.search(r'[A-Z][0-9]+\:[A-Z][0-9]+', cell._value)
						store_header = update2(cell, refrange.group(0), store_header)
					# if (token.value == "*"):
					# 	value = cell.value[1:]
					# 	sub_cell1 = sheet[value.split("*")[0]]
					# 	sub_cell2 = sheet[value.split("*")[1]]
					# 	store_header = update2(sub_cell1, cell, store_header)
					# 	store_header = update2(sub_cell2, cell, store_header)

						# if (token.value == ""): # refer to other cell
						# 	refcell =
						# 	store_header = update1(cell, refcell, store_header)
	# {A5:A2, A2:A1} 
	# print store_header
	for key, value in store_header.iteritems():
		if (headerchecker.check_and(key, store_header)):
			continue
		else:
			problem_cell.append(key)
	redFill = PatternFill(start_color = 'FFFF0000', end_color = 'FFFF0000', fill_type = 'solid')
	print ("Proble Cells:")

	for key, value in store_header.iteritems():
		for t in value:
			if cell.coordinate in t:
				problem_cell.append(key)
	for e in problem_cell:
		print e
		marked_cell = sheet.cell(e)
		marked_cell.fill = redFill
							

	wb.save("res.xlsx")








				
